"""Генератор счетов на оплату в формате Excel"""
from io import BytesIO
from decimal import Decimal
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage

from app.models import Contract, CLIENT_TYPES
from .constants import INVOICE_TEMPLATE_PATH, MONTHS_RU
from .utils import number_to_words_ru
from .replacements import get_short_name
from .qr_generator import generate_payment_qr_image


def build_client_invoice_line(client) -> str:
    """Формирует строку клиента для счета в зависимости от типа.

    Форматы:
    - ИП: ИП Фамилия И.О., ИНН xxx, ОГРНИП xxx
    - ООО/АО/ПАО/НКО: ООО «Название», ИНН xxx, ОГРН xxx, КПП xxx
    - ФЛ: Фамилия И.О., ИНН xxx
    """
    ct = client.client_type or "ip"

    short_name = get_short_name(client)

    if ct == "ip":
        parts = [f"ИП {short_name}"]
        if client.inn:
            parts.append(f"ИНН {client.inn}")
        if client.ogrn:
            parts.append(f"ОГРНИП {client.ogrn}")
    elif ct in ("ooo", "ao", "pao", "nko"):
        prefix = CLIENT_TYPES.get(ct, "")
        company = client.company_name or client.name
        parts = [f"{prefix} «{company}»"]
        if client.inn:
            parts.append(f"ИНН {client.inn}")
        if client.ogrn:
            parts.append(f"ОГРН {client.ogrn}")
        if client.kpp:
            parts.append(f"КПП {client.kpp}")
    else:  # fl - физлицо
        parts = [short_name]
        if client.inn:
            parts.append(f"ИНН {client.inn}")

    return ", ".join(parts)


def format_invoice_date(d) -> str:
    """Форматирует дату для счета: '5 января 2026 г.'"""
    return f"{d.day} {MONTHS_RU[d.month]} {d.year} г."


def format_price(amount: Decimal) -> str:
    """Форматирует сумму: '20 000,00'"""
    return f"{amount:,.2f}".replace(",", " ").replace(".", ",")


def format_price_words(amount: Decimal) -> str:
    """Форматирует сумму прописью: 'Двадцать тысяч рублей 00 копеек'"""
    rubles = int(amount)
    kopecks = int((amount - rubles) * 100)
    words = number_to_words_ru(Decimal(rubles))
    # Capitalize first letter
    words = words[0].upper() + words[1:]
    return f"{words} {kopecks:02d} копеек"


def replace_in_cell(cell, replacements: dict):
    """Заменяет плейсхолдеры в ячейке"""
    if cell.value is None:
        return

    value = str(cell.value)
    for key, replacement in replacements.items():
        if key in value:
            value = value.replace(key, str(replacement or ""))

    cell.value = value


def copy_row_style(ws, source_row: int, target_row: int, max_col: int = 70):
    """Копирует стиль строки"""
    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)


def fill_services_table(ws, services: list, start_row: int = 25) -> Decimal:
    """Заполняет таблицу услуг, возвращает итоговую сумму.

    Структура строки услуги (из шаблона):
    - B: № (номер)
    - D: Наименование услуги
    - Y: Код (00-00000006)
    - AJ: Кол-во (1)
    - AP: Ед. (шт)
    - AS: Цена
    - BC: Сумма
    """
    total = Decimal("0")

    # Если услуг больше одной, нужно вставить дополнительные строки
    if len(services) > 1:
        # Вставляем строки после start_row
        ws.insert_rows(start_row + 1, len(services) - 1)
        # Копируем стиль из первой строки
        for i in range(1, len(services)):
            copy_row_style(ws, start_row, start_row + i)

    for i, service in enumerate(services):
        row = start_row + i

        # № (колонка B = 2)
        ws.cell(row=row, column=2).value = i + 1

        # Наименование (колонка D = 4)
        ws.cell(row=row, column=4).value = service.name

        # Код (колонка Y = 25) - генерируем код услуги
        ws.cell(row=row, column=25).value = f"00-{service.id:08d}"

        # Кол-во (колонка AJ = 36)
        ws.cell(row=row, column=36).value = 1

        # Ед. (колонка AP = 42)
        ws.cell(row=row, column=42).value = "шт"

        # Цена (колонка AS = 45)
        ws.cell(row=row, column=45).value = float(service.price)

        # Сумма (колонка BC = 55)
        ws.cell(row=row, column=55).value = float(service.price)

        total += service.price

    return total


def update_totals(ws, total: Decimal, services_count: int, services_end_row: int):
    """Обновляет итоговые ячейки.

    Смещение строк зависит от количества услуг:
    - Базовые строки: 26 (итого кол-во), 28 (Итого), 29 (НДС), 30 (Всего к оплате)
    - Строки 32, 33 - текстовые итоги
    """
    # Вычисляем смещение строк
    offset = services_end_row - 25  # 25 - базовая строка услуги

    # Кол-во итого (строка 26 -> AJ26)
    ws.cell(row=26 + offset, column=36).value = services_count

    # Сумма итого (строка 26 -> BF26)
    ws.cell(row=26 + offset, column=58).value = float(total)

    # Итого (строка 28 -> BC28)
    ws.cell(row=28 + offset, column=55).value = float(total)

    # Всего к оплате (строка 30 -> BC30)
    ws.cell(row=30 + offset, column=55).value = float(total)

    # Текстовые итоги (строка 32, 33)
    ws.cell(row=32 + offset, column=2).value = (
        f"Всего наименований {services_count}, на сумму {format_price(total)} руб."
    )
    ws.cell(row=33 + offset, column=2).value = format_price_words(total)


def insert_payment_qr(
    ws,
    invoice_number: str,
    invoice_date: str,
    amount: Decimal,
    services_end_row: int,
):
    """Вставляет QR-код для оплаты в счет.

    Args:
        ws: Worksheet объект
        invoice_number: Номер счета
        invoice_date: Дата счета (DD.MM.YYYY)
        amount: Сумма к оплате
        services_end_row: Последняя строка таблицы услуг
    """
    # Генерируем QR-код
    qr_buffer = generate_payment_qr_image(
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        amount=amount,
        box_size=4,
        border=2,
    )

    # Создаем объект изображения для openpyxl
    qr_image = XLImage(qr_buffer)

    # Устанавливаем размер (~2.5 см x 2.5 см)
    qr_image.width = 95
    qr_image.height = 95

    # Вычисляем позицию: в самом низу документа, после подписи
    # Базовая строка подписи (М.П.): 45
    # Смещение: services_end_row - 25
    offset = services_end_row - 25
    qr_row = 50 + offset  # После подписи с отступом

    # Вставляем в ячейку B
    cell_anchor = f"B{qr_row}"
    ws.add_image(qr_image, cell_anchor)


def generate_invoice(contract: Contract) -> bytes:
    """Генерирует счет на оплату в формате Excel.

    Args:
        contract: Договор с услугами и клиентом

    Returns:
        bytes: Содержимое Excel файла
    """
    if not INVOICE_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Шаблон счета не найден: {INVOICE_TEMPLATE_PATH}")

    wb = load_workbook(INVOICE_TEMPLATE_PATH)
    ws = wb.active

    client = contract.client
    d = contract.date

    # Словарь замен для простых ячеек
    replacements = {
        "{{contract_number}}": contract.number,
        "{{contract_date}}": d.strftime("%d.%m.%Y"),
        "{{invoice_date}}": format_invoice_date(d),
        "{{client_invoice_line}}": build_client_invoice_line(client),
    }

    # Заменяем плейсхолдеры в заголовке и информации о клиенте
    # B10 - заголовок счета
    replace_in_cell(ws['B10'], replacements)
    # F17 - информация о клиенте
    replace_in_cell(ws['F17'], replacements)
    # F20 - основание (ссылка на договор)
    replace_in_cell(ws['F20'], replacements)

    # Заполняем таблицу услуг
    total = fill_services_table(ws, contract.services)
    services_count = len(contract.services)
    services_end_row = 25 + services_count - 1

    # Обновляем итоги
    update_totals(ws, total, services_count, services_end_row)

    # Добавляем QR-код для оплаты
    insert_payment_qr(
        ws=ws,
        invoice_number=contract.number,
        invoice_date=d.strftime("%d.%m.%Y"),
        amount=total,
        services_end_row=services_end_row,
    )

    # Сохраняем в байты
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
